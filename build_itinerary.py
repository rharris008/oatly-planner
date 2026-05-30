#!/usr/bin/env python3
"""Build the family holiday flights + accommodation workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# DATA  (2026 market-rate ESTIMATES in AUD; live booking sites are JS-rendered
#        and do not return live quotes, so figures are grounded in route/fare
#        research and clearly labelled as estimates. Party = 2 adults + 2 kids
#        aged 15 = 4 passengers, all on adult fares.)
# ----------------------------------------------------------------------------

# Sheet 1: Flights -- [Leg, Date, Route, Airline, Stops, Duration, $/pax, $x4, URL]
flights = [
    # LEG 1 -- BNE -> IST, 29 Nov, Premium Economy
    ["1. BNE→IST (Prem Econ)", "29 Nov 2026", "BNE-SIN-IST", "Singapore Airlines (A350, new Prem Econ from 1 Nov 2026)", "1 (Singapore SIN)", "~21h 20m", 3950, 15800,
     "https://www.google.com/travel/flights?q=Flights%20from%20BNE%20to%20IST%20on%202026-11-29%20premium%20economy"],
    ["1. BNE→IST (Prem Econ)", "29 Nov 2026", "BNE-DOH-IST", "Qatar Airways (QR899)", "1 (Doha DOH)", "~20h 45m", 3800, 15200,
     "https://www.google.com/travel/flights?q=Flights%20from%20BNE%20to%20IST%20on%202026-11-29%20premium%20economy"],
    ["1. BNE→IST (Prem Econ)", "29 Nov 2026", "BNE-DXB-IST", "Emirates (EK)", "1 (Dubai DXB)", "~22h 00m", 4100, 16400,
     "https://www.emirates.com/au/english/destinations/bne/ist/flights-from-brisbane-to-istanbul-airport/"],

    # LEG 2 -- IST -> LHR, ~5 Dec, Economy
    ["2. IST→LHR (Economy)", "5 Dec 2026", "IST-LHR (nonstop)", "Turkish Airlines (TK, B787)", "Nonstop", "~4h 15m", 240, 960,
     "https://www.google.com/travel/flights?q=Flights%20from%20IST%20to%20LHR%20on%202026-12-05%20economy"],
    ["2. IST→LHR (Economy)", "5 Dec 2026", "IST-LHR (nonstop)", "British Airways (BA, B787)", "Nonstop", "~4h 25m", 290, 1160,
     "https://www.google.com/travel/flights?q=Flights%20from%20IST%20to%20LHR%20on%202026-12-05%20economy"],
    ["2. IST→LHR (Economy)", "5 Dec 2026", "SAW-LGW (alt. airports)", "Pegasus Airlines (PC, ex Sabiha Gökçen)", "Nonstop", "~4h 20m", 130, 520,
     "https://www.google.com/travel/flights?q=Flights%20from%20SAW%20to%20London%20on%202026-12-05%20economy"],

    # LEG 3 -- LHR -> GVA, ~16 Dec, Economy
    ["3. LHR→GVA (Economy)", "16 Dec 2026", "LHR-GVA (nonstop)", "British Airways (BA mainline)", "Nonstop", "~1h 45m", 340, 1360,
     "https://www.britishairways.com/travel/fx/public/en_gb?from=LHR&to=GVA&depDate=2026-12-16&ADT=2&CHD=2"],
    ["3. LHR→GVA (Economy)", "16 Dec 2026", "LHR-GVA (nonstop)", "SWISS (LX, Lufthansa Group)", "Nonstop", "~1h 40m", 365, 1460,
     "https://www.swiss.com/gb/en/book/flights/LHR/GVA?date=2026-12-16&adults=2&children=2"],
    ["3. LHR→GVA (Economy)", "16 Dec 2026", "LGW-GVA (Gatwick, not LHR)", "easyJet (U2)", "Nonstop", "~1h 35m", 210, 840,
     "https://www.easyjet.com/en/cheap-flights/london-gatwick/geneva"],

    # LEG 4 -- GVA -> CNS, ~26 Dec, Economy, any hub (Christmas peak)
    ["4. GVA→CNS (Economy)", "26 Dec 2026", "GVA-SIN-CNS", "Singapore Airlines (only carrier flying int'l into CNS)", "1 (Singapore SIN)", "~24–26h", 2650, 10600,
     "https://www.google.com/travel/flights?q=Flights%20from%20GVA%20to%20CNS%20on%202026-12-26"],
    ["4. GVA→CNS (Economy)", "26 Dec 2026", "GVA-DXB-BNE-CNS", "Emirates + Qantas/Virgin (BNE-CNS)", "2 (Dubai DXB, Brisbane BNE)", "~30–33h", 2500, 10000,
     "https://www.google.com/travel/flights?q=Flights%20from%20GVA%20to%20CNS%20on%202026-12-26"],
    ["4. GVA→CNS (Economy)", "26 Dec 2026", "GVA-DOH-BNE-CNS", "Qatar Airways + Qantas/Virgin domestic", "2 (Doha DOH, Brisbane BNE)", "~31–34h", 2400, 9600,
     "https://www.google.com/travel/flights?q=Flights%20from%20GVA%20to%20CNS%20on%202026-12-26"],

    # LEG 5 -- CNS -> OOL, ~31 Dec, Economy
    ["5. CNS→OOL (Economy)", "31 Dec 2026", "CNS-OOL (nonstop)", "Jetstar (JQ) — only nonstop", "Nonstop", "~2h 20m", 220, 880,
     "https://www.jetstar.com/au/en/home"],
    ["5. CNS→OOL (Economy)", "31 Dec 2026", "CNS-BNE-OOL", "Virgin Australia (VA)", "1 (Brisbane BNE)", "~4–6h", 300, 1200,
     "https://www.virginaustralia.com/au/en/"],
    ["5. CNS→OOL (Economy)", "31 Dec 2026", "CNS-BNE-OOL", "Qantas (QF / QantasLink)", "1 (Brisbane BNE)", "~4–6h", 340, 1360,
     "https://www.qantas.com/au/en.html"],
]

# Sheet 2: Accommodation -- [Destination, Hotel, Stars, Room Type, Nights, $/night, $total, URL]
accom = [
    # ISTANBUL -- 29 Nov -> 4 Dec, 5 nights
    ["Istanbul (Sultanahmet)\n29 Nov–4 Dec (5 nts)", "Hotel Sari Konak", "4★", "Family/Triple Suite (sleeps 4)", 5, 165, 825,
     "https://www.booking.com/hotel/tr/sari-konak.html?checkin=2026-11-29&checkout=2026-12-04&group_adults=2&group_children=2&age=15&age=15"],
    ["Istanbul (Sultanahmet)\n29 Nov–4 Dec (5 nts)", "Mitspark Hotel", "4★", "Family Room (sleeps 4)", 5, 210, 1050,
     "https://www.booking.com/hotel/tr/mitspark.html?checkin=2026-11-29&checkout=2026-12-04&group_adults=2&group_children=2&age=15&age=15"],
    ["Istanbul (Sultanahmet)\n29 Nov–4 Dec (5 nts)", "Seven Hills Palace & Spa", "4★", "Family Suite / 2 connecting rooms", 5, 230, 1150,
     "https://www.booking.com/hotel/tr/seven-hills-palace.html?checkin=2026-11-29&checkout=2026-12-04&group_adults=2&group_children=2&age=15&age=15"],

    # CAPPADOCIA (Goreme) -- 4 Dec -> 7 Dec, 3 nights
    ["Cappadocia (Göreme)\n4–7 Dec (3 nts)", "Cave Hotel Saksagan", "4★ boutique cave", "Family Cave Room (sleeps 4)", 3, 130, 390,
     "https://www.booking.com/hotel/tr/cave-saksagan.html?checkin=2026-12-04&checkout=2026-12-07&group_adults=2&group_children=2&age=15&age=15"],
    ["Cappadocia (Göreme)\n4–7 Dec (3 nts)", "Aren Cave Hotel & Art Gallery", "4★ boutique cave", "Family Cave Suite (sleeps 4)", 3, 170, 510,
     "https://www.booking.com/hotel/tr/aren-cave.html?checkin=2026-12-04&checkout=2026-12-07&group_adults=2&group_children=2&age=15&age=15"],
    ["Cappadocia (Göreme)\n4–7 Dec (3 nts)", "Zara Cave Hotel", "4★ boutique cave", "Family Cave Room (sleeps 4)", 3, 185, 555,
     "https://www.booking.com/hotel/tr/zara-cave.html?checkin=2026-12-04&checkout=2026-12-07&group_adults=2&group_children=2&age=15&age=15"],

    # CHAMONIX -- 17 Dec -> 25 Dec, 8 nights (Christmas peak)
    ["Chamonix\n17–25 Dec (8 nts)", "Hotel L'Oustalet", "4★", "1 Family Room (2A+2C)", 8, 467, 3735,
     "https://www.booking.com/hotel/fr/oustalet.html?checkin=2026-12-17&checkout=2026-12-25&group_adults=2&group_children=2&age=15&age=15"],
    ["Chamonix\n17–25 Dec (8 nts)", "Auberge du Manoir", "3★", "2 interconnecting/adjacent rooms", 8, 520, 4160,
     "https://www.booking.com/hotel/fr/auberge-du-manoir.html?checkin=2026-12-17&checkout=2026-12-25&group_adults=2&group_children=2&age=15&age=15"],
    ["Chamonix\n17–25 Dec (8 nts)", "RockyPop Chamonix – Les Houches", "3★", "1 Family Room (2A+2C)", 8, 340, 2720,
     "https://www.booking.com/hotel/fr/rockypop-aux-portes-de-chamonix.html?checkin=2026-12-17&checkout=2026-12-25&group_adults=2&group_children=2&age=15&age=15"],

    # CAIRNS -- 26 Dec -> 31 Dec, 5 nights, 2 rooms (NYE peak)
    ["Cairns\n26–31 Dec (5 nts)", "Rydges Esplanade Resort Cairns", "4★", "2 × Resort/Hotel rooms", 5, 230, 2300,
     "https://www.booking.com/hotel/au/rydges-esplanade-resort-cairns.html?checkin=2026-12-26&checkout=2026-12-31&group_adults=2&group_children=0&no_rooms=2"],
    ["Cairns\n26–31 Dec (5 nts)", "Mantra Esplanade Cairns", "4★", "2 × Hotel rooms", 5, 250, 2500,
     "https://www.booking.com/hotel/au/mantra-esplanade.html?checkin=2026-12-26&checkout=2026-12-31&group_adults=2&group_children=0&no_rooms=2"],
    ["Cairns\n26–31 Dec (5 nts)", "Pacific Hotel Cairns", "4★", "2 × Superior rooms", 5, 215, 2150,
     "https://www.booking.com/hotel/au/pacific-international-cairns.html?checkin=2026-12-26&checkout=2026-12-31&group_adults=2&group_children=0&no_rooms=2"],
]

# Cost summary -- recommended (cheapest sensible) pick per leg/destination
flight_picks = [
    ("1. BNE→IST (Prem Econ)", "Qatar Airways (BNE-DOH-IST)", 15200),
    ("2. IST→LHR (Economy)", "Turkish Airlines (nonstop IST-LHR)", 960),
    ("3. LHR→GVA (Economy)", "British Airways (nonstop, true LHR)", 1360),
    ("4. GVA→CNS (Economy)", "Qatar Airways via DOH/BNE", 9600),
    ("5. CNS→OOL (Economy)", "Jetstar (nonstop)", 880),
]
accom_picks = [
    ("Istanbul (Sultanahmet, 5 nts)", "Mitspark Hotel (4★)", 1050),
    ("Cappadocia / Göreme (3 nts)", "Aren Cave Hotel (4★)", 510),
    ("Chamonix (8 nts)", "RockyPop Chamonix (3★)", 2720),
    ("Cairns (5 nts, 2 rooms)", "Pacific Hotel Cairns (4★)", 2150),
]

# ----------------------------------------------------------------------------
# STYLING HELPERS
# ----------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
TOTAL_FILL = PatternFill("solid", fgColor="C6E0B4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
LINK_FONT = Font(color="0563C1", underline="single", size=10)
BOLD = Font(bold=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def aud(cell):
    cell.number_format = '"A$"#,##0'


wb = Workbook()

# ============================ SHEET 1: FLIGHTS ==============================
ws = wb.active
ws.title = "Flights"
ws["A1"] = "Family Holiday — Flights (top 3 options per leg)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Party: 2 adults + 2 children (age 15) = 4 passengers · Prices are 2026 market-rate ESTIMATES in AUD · age 15 = adult fare"
ws["A2"].font = Font(italic=True, size=9, color="808080")

headers = ["Leg", "Date", "Route", "Airline", "Stops", "Duration", "Price/Person (AUD)", "Total ×4 (AUD)", "Booking URL"]
hrow = 4
for c, h in enumerate(headers, 1):
    ws.cell(row=hrow, column=c, value=h)
style_header(ws, hrow, len(headers))

r = hrow + 1
prev_leg = None
for row in flights:
    leg, date, route, airline, stops, dur, ppax, tot, url = row
    ws.cell(row=r, column=1, value=leg).alignment = WRAP_TOP
    ws.cell(row=r, column=2, value=date).alignment = WRAP_TOP
    ws.cell(row=r, column=3, value=route).alignment = WRAP_TOP
    ws.cell(row=r, column=4, value=airline).alignment = WRAP_TOP
    ws.cell(row=r, column=5, value=stops).alignment = WRAP_TOP
    ws.cell(row=r, column=6, value=dur).alignment = WRAP_TOP
    cp = ws.cell(row=r, column=7, value=ppax); aud(cp); cp.alignment = CENTER
    ct = ws.cell(row=r, column=8, value=tot); aud(ct); ct.alignment = CENTER; ct.font = BOLD
    lc = ws.cell(row=r, column=9, value="Open booking link"); lc.hyperlink = url; lc.font = LINK_FONT
    lc.alignment = WRAP_TOP
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
    # shade alternating leg groups
    if leg != prev_leg:
        prev_leg = leg
        band = (flights.index(row) // 3) % 2 == 1
    if (flights.index(row) // 3) % 2 == 1:
        for c in range(1, len(headers) + 1):
            if not ws.cell(row=r, column=c).fill.fgColor.rgb or ws.cell(row=r, column=c).fill.patternType is None:
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F2F2F2")
    r += 1

widths = [22, 12, 24, 34, 20, 12, 16, 14, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ========================= SHEET 2: ACCOMMODATION ==========================
ws2 = wb.create_sheet("Accommodation")
ws2["A1"] = "Family Holiday — Accommodation (top 3 options per destination)"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = "Family of 4 (2 adults + 2 children age 15) · Booking.com · Prices are 2026 market-rate ESTIMATES in AUD"
ws2["A2"].font = Font(italic=True, size=9, color="808080")

headers2 = ["Destination / Dates", "Hotel Name", "Star Rating", "Room Type", "Nights", "Price/Night (AUD)", "Total Stay (AUD)", "Booking URL"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=c, value=h)
style_header(ws2, 4, len(headers2))

r = 5
for row in accom:
    dest, hotel, stars, rtype, nights, pn, tot, url = row
    ws2.cell(row=r, column=1, value=dest).alignment = WRAP_TOP
    ws2.cell(row=r, column=2, value=hotel).alignment = WRAP_TOP
    ws2.cell(row=r, column=3, value=stars).alignment = CENTER
    ws2.cell(row=r, column=4, value=rtype).alignment = WRAP_TOP
    ws2.cell(row=r, column=5, value=nights).alignment = CENTER
    cp = ws2.cell(row=r, column=6, value=pn); aud(cp); cp.alignment = CENTER
    ct = ws2.cell(row=r, column=7, value=tot); aud(ct); ct.alignment = CENTER; ct.font = BOLD
    lc = ws2.cell(row=r, column=8, value="Open booking link"); lc.hyperlink = url; lc.font = LINK_FONT
    lc.alignment = WRAP_TOP
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=r, column=c).border = BORDER
    if (accom.index(row) // 3) % 2 == 1:
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F2F2F2")
    r += 1

widths2 = [26, 32, 16, 30, 8, 16, 16, 20]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A5"

# ============================== SHEET 3: COST ==============================
ws3 = wb.create_sheet("Cost")
ws3["A1"] = "Family Holiday — Cost Summary (recommended best-value picks)"
ws3["A1"].font = TITLE_FONT
ws3["A2"] = "Party: 2 adults + 2 children (15). All figures AUD, 2026 market-rate ESTIMATES."
ws3["A2"].font = Font(italic=True, size=9, color="808080")

r = 4
# --- Flights block ---
ws3.cell(row=r, column=1, value="FLIGHTS").font = BOLD
ws3.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 4):
    ws3.cell(row=r, column=c).fill = SECTION_FILL
r += 1
for c, h in enumerate(["Leg", "Recommended option", "Total ×4 (AUD)"], 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 3)
r += 1
flight_total = 0
flight_start = r
for leg, opt, cost in flight_picks:
    ws3.cell(row=r, column=1, value=leg).alignment = WRAP_TOP
    ws3.cell(row=r, column=2, value=opt).alignment = WRAP_TOP
    cc = ws3.cell(row=r, column=3, value=cost); aud(cc); cc.alignment = CENTER
    for c in range(1, 4):
        ws3.cell(row=r, column=c).border = BORDER
    flight_total += cost
    r += 1
ws3.cell(row=r, column=2, value="Flights subtotal").font = BOLD
fc = ws3.cell(row=r, column=3, value=f"=SUM(C{flight_start}:C{r-1})"); aud(fc); fc.font = BOLD; fc.alignment = CENTER
fc.fill = TOTAL_FILL; ws3.cell(row=r, column=2).fill = TOTAL_FILL
flights_subtotal_row = r
r += 2

# --- Accommodation block ---
ws3.cell(row=r, column=1, value="ACCOMMODATION").font = BOLD
for c in range(1, 4):
    ws3.cell(row=r, column=c).fill = SECTION_FILL
r += 1
for c, h in enumerate(["Destination", "Recommended option", "Total Stay (AUD)"], 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 3)
r += 1
accom_start = r
for dest, opt, cost in accom_picks:
    ws3.cell(row=r, column=1, value=dest).alignment = WRAP_TOP
    ws3.cell(row=r, column=2, value=opt).alignment = WRAP_TOP
    cc = ws3.cell(row=r, column=3, value=cost); aud(cc); cc.alignment = CENTER
    for c in range(1, 4):
        ws3.cell(row=r, column=c).border = BORDER
    r += 1
ws3.cell(row=r, column=2, value="Accommodation subtotal").font = BOLD
ac = ws3.cell(row=r, column=3, value=f"=SUM(C{accom_start}:C{r-1})"); aud(ac); ac.font = BOLD; ac.alignment = CENTER
ac.fill = TOTAL_FILL; ws3.cell(row=r, column=2).fill = TOTAL_FILL
accom_subtotal_row = r
r += 2

# --- Grand total ---
ws3.cell(row=r, column=2, value="GRAND TOTAL (flights + accommodation)").font = Font(bold=True, size=12)
gc = ws3.cell(row=r, column=3, value=f"=C{flights_subtotal_row}+C{accom_subtotal_row}")
aud(gc); gc.font = Font(bold=True, size=12); gc.alignment = CENTER
for c in range(1, 4):
    ws3.cell(row=r, column=c).fill = TOTAL_FILL
    ws3.cell(row=r, column=c).border = BORDER
grand_row = r
r += 2

# --- Notes ---
notes = [
    "Notes & assumptions:",
    "• Excludes: travel between airports/hotels, local transport, ski passes/lift tickets, meals, travel insurance, visas (Türkiye e-visa, UK/Schengen as applicable), and the BNE→home final domestic leg.",
    "• Long-haul GVA→CNS (26 Dec) sits in the deepest Christmas/New-Year peak — the single biggest cost; flexing dates off 26 Dec could save A$400–700/person.",
    "• Children aged 15 are charged adult fares by virtually all airlines.",
    "• Prices are market-rate ESTIMATES (live booking engines are JS-rendered and don't return live quotes); use the booking links on the Flights/Accommodation sheets to pull live prices.",
    "• 'Recommended option' = best-value sensible pick per leg/destination; cheaper alt-airport options (e.g. Pegasus SAW-LGW, easyJet LGW-GVA) are listed on the Flights sheet.",
]
for n in notes:
    cell = ws3.cell(row=r, column=1, value=n)
    cell.font = BOLD if n.endswith(":") else Font(size=9, color="595959")
    cell.alignment = WRAP_TOP
    r += 1

ws3.column_dimensions["A"].width = 34
ws3.column_dimensions["B"].width = 44
ws3.column_dimensions["C"].width = 18

# ----------------------------------------------------------------------------
out = "Family_Holiday_Itinerary_2026.xlsx"
wb.save(out)
print(f"Saved {out}")
print(f"Flights subtotal (recommended): A${flight_total:,}")
print(f"Accommodation subtotal (recommended): A${sum(c for _,_,c in accom_picks):,}")
print(f"Grand total: A${flight_total + sum(c for _,_,c in accom_picks):,}")
